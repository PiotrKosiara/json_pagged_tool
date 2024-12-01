import os
import json
import openpyxl
from openpyxl.styles import PatternFill


# Ładowanie konfiguracji z pliku config.json
def load_config(config_file="config.json"):
    with open(config_file, "r", encoding="utf-8") as f:
        return json.load(f)


# Funkcja do spłaszczania zagnieżdżonych słowników
def flatten_json(data, parent_key=""):
    """
    Rekurencyjnie spłaszcza słownik, tworząc klucze o pełnej ścieżce.
    """
    items = []
    for key, value in data.items():
        new_key = f"{parent_key}.{key}" if parent_key else key
        if isinstance(value, dict):
            items.extend(flatten_json(value, new_key).items())
        elif isinstance(value, list):
            for idx, item in enumerate(value):
                items.extend(flatten_json(item, f"{new_key}[{idx}]").items())
        else:
            items.append((new_key, value))
    return dict(items)


# Funkcja do porównywania dwóch spłaszczonych JSON-ów
def compare_json(final, test):
    """
    Porównuje dwie struktury JSON i zwraca różnice na poziomie najbardziej wewnętrznych kluczy.
    """
    flattened_final = flatten_json(final)
    flattened_test = flatten_json(test)

    comparison = {}
    for key in flattened_final:
        if key in flattened_test:
            comparison[key] = 1 if flattened_final[key] == flattened_test[key] else 0
        else:
            comparison[key] = 0
    for key in flattened_test:
        if key not in flattened_final:
            comparison[key] = 0
    return comparison


# Główna funkcja do generowania raportów
def generate_report(config):
    golden_dir = config["golden_dir"]
    test_dir = config["test_dir"]
    output_file = config["output_file"]

    # Kolory do raportu
    green_fill = PatternFill(**config["styles"]["green"])
    red_fill = PatternFill(**config["styles"]["red"])

    # Tworzenie arkusza Excel
    workbook = openpyxl.Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "Details"
    detail_sheet.append(["No", "Golden File", "Test File", "Key", "Match"])

    summary_sheet = workbook.create_sheet(title="Summary")
    summary_sheet.append(["No", "Golden File", "Test File", "Accuracy (%)"])

    # Przetwarzanie plików JSON
    files = os.listdir(golden_dir)
    no = 1
    for file in files:
        if "_final" in file:
            golden_path = os.path.join(golden_dir, file)
            test_file = file.replace("_final", "_test")
            test_path = os.path.join(test_dir, test_file)

            if os.path.exists(test_path):
                with open(golden_path, "r", encoding="utf-8") as f:
                    final_data = json.load(f)
                with open(test_path, "r", encoding="utf-8") as f:
                    test_data = json.load(f)

                comparison = compare_json(final_data, test_data)
                matches = 0
                total = len(comparison)
                for key, match in comparison.items():
                    row = [no, file, test_file, key, match]
                    detail_sheet.append(row)
                    matches += match

                    # Kolorowanie komórek
                    match_cell = detail_sheet.cell(row=detail_sheet.max_row, column=5)
                    match_cell.fill = green_fill if match == 1 else red_fill

                accuracy = (matches / total) * 100 if total > 0 else 0
                summary_sheet.append([no, file, test_file, round(accuracy, 2)])
            else:
                print(f"Test file not found: {test_file}")
                detail_sheet.append([no, file, test_file, "File missing", 0])
                summary_sheet.append([no, file, test_file, 0])

            no += 1

    # Zapisywanie raportu do pliku
    workbook.save(output_file)
    print(f"Report generated: {output_file}")


# Uruchamianie skryptu
if __name__ == "__main__":
    config = load_config()
    generate_report(config)


